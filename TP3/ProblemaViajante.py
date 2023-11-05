'''ENUNCIADO 3ER TRABAJO PRÁCTICO
El problema del viajante (también conocido como problema del viajante de comercio o por sus siglas en inglés: TSP (Traveling Salesman Problem), 
es uno de los problemas más famosos (y quizás el mejor estudiado) en el campo de la optimización combinatoria computacional.
A pesar de la aparente sencillez de su planteamiento, el TSP es uno de los más complejos de resolver.

Definición: Sean N ciudades de un territorio. La distancia entre cada ciudad viene dada por la matriz D: NxN, donde d[x,y] representa la distancia que hay entre la ciudad X y la ciudad Y.
El objetivo es encontrar una ruta que, comenzando y terminando en una ciudad concreta, pase una sola vez por cada una de las ciudades y minimice la distancia recorrida por el viajante.

Ejercicios:
1.Hallar la ruta de distancia mínima que logre unir todas las capitales de provincias de la República Argentina, utilizando un método exhaustivo. ¿Puede resolver el problema? 
Justificar de manera teórica.
2.Realizar un programa que cuente con un menú con las siguientes opciones:
    a)Permitir ingresar una provincia y hallar la ruta de distancia mínima que logre unir todas las capitales de provincias de la República Argentina 
    partiendo de dicha capital utilizando la siguiente heurística: “Desde cada ciudad ir a la ciudad más cercana no visitada.”  Recordar regresar siempre a la ciudad de partida. 
    Presentar un mapa de la República con el recorrido indicado. Además indicar la ciudad de partida, el recorrido completo y la longitud del trayecto. 
    El programa deberá permitir seleccionar la capital que el usuario desee ingresar como inicio del recorrido.
    b)Encontrar el recorrido mínimo para visitar todas las capitales de las provincias de la República Argentina siguiendo la heurística mencionada en el punto a. 
    Deberá mostrar como salida el recorrido y la longitud del trayecto.
    c)Hallar la ruta de distancia mínima que logre unir todas las capitales de provincias de la República Argentina, utilizando un algoritmo genético.

Recomendaciones para el algoritmo:
N = 50 Número de cromosomas de las poblaciones.
M = 200 Cantidad de ciclos.
Cromosomas: permutaciones de 23 números naturales del 1 al 23 donde cada gen es una ciudad.
Las frecuencias de crossover y de mutación quedan a criterio del grupo.
Se deberá usar crossover cíclico.
Comparar los resultados obtenidos  entre la resolución a través de heurísticas y con algoritmos genéticos a través de una conclusión que deberá anexarse al informe.
Agregar en el informe un apartado final denominado «Aportes Prácticos del TSP» donde se expliquen algunas aplicaciones en las que actualmente se use el problema del viajante. 
Tomar por lo menos dos y explicarlas.'''

from os import system
import os 
import numpy as np
import openpyxl as Excel
import random as r
import GeneticoCrossOver as CrossOver
import GeneticoRuleta as Ruleta
import GeneticoFitness as Fitness
import Geneticotorneo as Torneo
import GeneticoGraficar as Graficar
import GeneticoElitismo as Elitismo
Ciudades_Disponibles = []
Distancia_Recorrida = 0
Descripcion_Recorrido = ''
Id_Ciudad_Actual = 0
Id_Ciudad_Cercana = 0

cwd             = os.getcwd()               # Creo que aca depende donde estamos apuntando el espacio de trabajo actualmente...
                                            # En mi caso, de una PC a otro me difiere el espacio de trabajo de TP3 a AG_2023/TP3
                                            # En caso de error en la consola aparece a donde esta apuntando el entorno
                                            # para sus pruebas editar el valor de la variable ExcelDocument en una de las primeras lineas del main() para que apunte bien
                                            # CONSIDERACIÓN: Quizas habría que investigar para encontra algo mejor como un proceso que devuelva la ubicacion del programa actual

FileName        = 'Distancias_Ruta.xlsx'    # Yo me arme esta tabla excel con distancias por ruta. si quieren usar la que brinda la catedra hay que meterla en la carpeta del proyecto y apuntar ahi
                                            # O quizas crear una solapa en el actual archivo xlsx y apuntar a esa solapa. Pineso que de esta manera hacemos un aporte a la materia (le regalamos el excel para futuras generaciones)                                          

class Constant : 
    CIUDADES = np.array((
        ['A','Córdoba'                               ],
        ['B','Corrientes'                            ],
        ['C','Formosa'                               ],
        ['D','La Plata'                              ],
        ['E','La Rioja'                              ],
        ['F','Mendoza'                               ],
        ['G','Neuquén'                               ],
        ['H','Paraná'                                ],
        ['I','Posadas'                               ],
        ['J','Rawson'                                ],
        ['K','Resistencia'                           ],
        ['L','Río Gallegos'                          ],
        ['M','San Fernando del Valle de Catamarca'   ],
        ['N','San Miguel de Tucumán'                 ],
        ['O','Salta'                                 ],
        ['P','San Juan'                              ],
        ['Q','San Luis'                              ],
        ['R','San Salvador de Jujuy'                 ],
        ['S','Santa Fe'                              ],
        ['T','Santa Rosa'                            ],
        ['U','Santiago del Estero'                   ],
        ['V','Ushuaia'                               ],
        ['W','Viedma'                                ]))
    ELIT = 2 
    POBLACION_INICIAL = 50

def main():
    global cwd
    global FileName
    global ExcelDocument
    global Sheet

    ExcelDocument   = Excel.load_workbook(cwd+'\TP3\ '.strip() +FileName)  # Strip es para quitar los espacios en blanco, python no me deja poner la barra invertida al final de la cadena
    Sheet           = ExcelDocument.get_sheet_by_name('PorRuta')
    
    OPC = 0 
    while OPC == 0:
        system("cls")
        print('Problema del viajante')
        print(' 1- Método Heurístico')
        print(' 2- Algorítmo Genético')
        print('')
        OPC = int(input('OPCIÓN: '))
        if OPC > 2 or OPC < 0:
            print('Opción Incorrecta, vuelva a intentar...')
            OPC = 0
    match OPC:
        case 1:
            OPC2 = 0
            while OPC2 == 0:
                system("cls")
                print('Método Heurístico')
                print(' 1 - Con ciudad inicial')
                print(' 2 - Sin ciudad inicial')
                print('')
                OPC2 = int(input('OPCIÓN: '))
                if OPC2 > 2 or OPC2 < 0:
                    print('Opción Incorrecta, vuelva a intentar...')
                    OPC2 = 0
            match OPC2:
                case 1:
                    OPC3 = 0
                    while OPC3 == 0:
                        system("cls")
                        print('Seleccione Ciudad Inicial:')
                        print('')
                        ContCiudad = 0
                        for Ciudad in Constant.CIUDADES:
                            ContCiudad+=1
                            print(str(ContCiudad)+' - '+Ciudad[1])
                        print('')
                        OPC3 = int(input('Seleccione Opción: '))
                        if OPC3 > 23 or OPC3 < 0:
                            print('Opción Incorrecta, vuelva a intentar...')
                            OPC3 = 0
                        else:
                            Heuristico_Con_Ciudad(OPC3)
                            ExcelDocument.save(Constant.CIUDADES[Id_Ciudad_Actual-1][1]+'_'+FileName)
                            ExcelDocument.close()
                case 2:
                    Heuristico_Sin_Ciudad()
                    ExcelDocument.save(Constant.CIUDADES[Id_Ciudad_Actual-1][1]+'_'+FileName)
                    ExcelDocument.close()
        case 2:
            Algoritmo_Genetico()
    Volver_Menu()

def Volver_Menu():
    OPC = 0
    while OPC == 0:
        print('¿ Desea volver al menú ?')
        print('    1-SI   2-NO')
        OPC = int(input('OPCIÓN: '))
        if OPC > 2 or OPC < 0:
            print('Opción Incorrecta, vuelva a intentar...')
            OPC = 0
    match OPC:
        case 1:
            main()
        case 2:
            system("cls")
            print('Ejecución Finalizada!')

def Heuristico_Con_Ciudad(Id_Ciudad_Seleccionada):
    system("cls")
    print('Heuristico_Con_Ciudad')
    print('')
    print('Id_Ciudad_Seleccionada = '+str(Id_Ciudad_Seleccionada)+' - '+Constant.CIUDADES[Id_Ciudad_Seleccionada-1][1])
    Heuristica(Id_Ciudad_Seleccionada)

def Heuristico_Sin_Ciudad():
    system("cls")
    print('Heuristico_Sin_Ciudad')
    print('')
    Id_Ciudad_Random = r.randint(1,23)
    
    print('Id_Ciudad_Random = '+str(Id_Ciudad_Random)+' - '+Constant.CIUDADES[Id_Ciudad_Random-1][1])
    Heuristica(Id_Ciudad_Random)    

def Heuristica(Id_Ciudad_Inicial):
    global Ciudades_Disponibles
    global Id_Ciudad_Actual
    global Id_Ciudad_Cercana
    global Descripcion_Recorrido
    global Distancia_Recorrida

    Distancia_Recorrida = 0
    Descripcion_Recorrido = ''
    my_red      = Excel.styles.colors.Color(rgb='00FF0000')
    my_green    = Excel.styles.colors.Color(rgb='0000FF00')
    fillRed     = Excel.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    fillGreen   = Excel.styles.fills.PatternFill(patternType='solid', fgColor=my_green)

    for i in range(23):
        Ciudades_Disponibles.append(i+1)
    Ciudades_Disponibles.remove(Id_Ciudad_Inicial)
    Distancia_Recorrida   = 0
    Descripcion_Recorrido = ''
    Id_Ciudad_Actual = Id_Ciudad_Inicial
    Id_Ciudad_Cercana = Ciudad_Cercana_Disponible(Id_Ciudad_Actual)
    for i in range(22):
        Distancia_Recorrida     += Sheet.cell(row=Id_Ciudad_Actual,column=Id_Ciudad_Cercana).value
        #print('PINTAR ROJO'+str(Id_Ciudad_Actual)+ '-' + str(Id_Ciudad_Cercana))
        Sheet.cell(row=Id_Ciudad_Actual,column=Id_Ciudad_Cercana).fill = fillRed
        Descripcion_Recorrido   += '('+Constant.CIUDADES[Id_Ciudad_Actual-1][0]+') ' + Constant.CIUDADES[Id_Ciudad_Actual-1][1] + ' -> '
        Id_Ciudad_Actual = Id_Ciudad_Cercana
        Id_Ciudad_Cercana = Ciudad_Cercana_Disponible(Id_Ciudad_Actual)
    Id_Ciudad_Cercana = Id_Ciudad_Inicial
    Ciudades_Disponibles.append(Id_Ciudad_Inicial)
    #print('Viajes de regreso------------------> '+str(Id_Ciudad_Cercana))
    #print(Ciudades_Disponibles)
    Distancia_Recorrida     += Sheet.cell(row=Id_Ciudad_Actual,column=Id_Ciudad_Cercana).value
    Sheet.cell(row=Id_Ciudad_Actual,column=Id_Ciudad_Cercana).fill = fillRed
    Descripcion_Recorrido   += '('+Constant.CIUDADES[Id_Ciudad_Actual-1][0]+') ' + Constant.CIUDADES[Id_Ciudad_Actual-1][1] + ' -> '
    Id_Ciudad_Actual = Id_Ciudad_Inicial
    Descripcion_Recorrido   += '('+Constant.CIUDADES[Id_Ciudad_Actual-1][0]+') ' + Constant.CIUDADES[Id_Ciudad_Actual-1][1]
    #print('PINTAR VERDE '+str(Id_Ciudad_Inicial)+ '-' + str(Id_Ciudad_Inicial))
    Sheet.cell(row=Id_Ciudad_Inicial,column=Id_Ciudad_Inicial).fill = fillGreen
    print('RECORRIDO: '+Descripcion_Recorrido)
    print('')
    print('TOTAL KMS Recorridos: '+str(Distancia_Recorrida)+' Kms')
    print('')

def Algoritmo_Genetico():
    system("cls")
    print('Algoritmo_Genetico')
    print('')
    # Acá a diferencia de los TPs anteriores, en donde usabamos 1 o 0, vamos a tener que orientar el algoritmo a cromosomas del tipo ABCDEFGHIJKLMNOPQRSTUVW 
    # (23 LETRAS, Una para cada ciduad) ya que es importante saber el orden de como visita las ciudades
    Cromosomas = []
    Cromosomas = Crear_Cromosomas_Iniciales(50)  # llamada a la funcion que crea poblacion inicial
    poblacion = Cromosomas
    maximos = []
    crom_max = []
    minimos = []
    promedios = []
    elit =[]
    # binario_COEF = np.binary_repr(Constant.COEF, 0)
    cant_Genes = 23  # largo de los genes
    #print(poblacion )
    for i in range(200):
        resultados = []  # lista de las funciones objetivos de la poblacion actual
        for j in range(len(poblacion)):
             resultados.append(Funcion_Objetivo(poblacion[j]))

         #calcula max, min y promedio
        crom_max.append(poblacion[resultados.index(np.max(resultados))])
        maximos.append(np.max(resultados))
        minimos.append(np.min(resultados))
        promedios.append(np.average(resultados))
        #print(poblacion)
        
    #     # poblacion, fitnae_pob, cant_elit ->elit
        elit= Elitismo.elitismo(poblacion, Fitness.Fitness( resultados), Constant.ELIT) 
    #     print(poblacion)
    #     # poblacion,fitnes_de_pob,cant_selecciones -> poblacionseleccionada
        
        poblacion = Ruleta.seleccion(poblacion, Fitness.Fitness(resultados), Constant.POBLACION_INICIAL - Constant.ELIT)
    #     poblacion = torneo.torneo(poblacion, Fitness.Fitness(resultados), Constant.POBLACION_INICIAL - Constant.ELIT)

    #     # pob_selec,const_cross,largo_gen -> poblacion_hijos
    
        poblacion = CrossOver.CrossOver( poblacion, 0.752, cant_Genes)     
        print() 
        #input()
    #     # poblacion,const_mut,largo_gen -> poblacion final
    #     poblacion = mutacion.mutar_poblacion(
    #         poblacion, Constant.P_MUTACION, cant_Genes)
        
    #     #Agrega a la poblacion seleccionada los cromosomas elites
        poblacion=poblacion+elit
        print(poblacion)

    Graficar.Tabla(crom_max,maximos,minimos,promedios)
    Graficar.MAX_MIN_PROM(maximos,minimos,promedios)   


def Ciudad_Cercana_Disponible(Ciudad_Inicial):
    global Ciudades_Disponibles
    Distancia_Menor = 10000000
    Ciudad_Mas_Cercana = 0
    #print('La ciduad Inicial es '+str(Ciudad_Inicial))
    for Ciudad in Ciudades_Disponibles:
        #print('La ciudad '+str(Ciudad)+' se encuentra disponible')
        Distancia = Sheet.cell(row = Ciudad_Inicial, column = Ciudad).value
        if Distancia < Distancia_Menor:
            Distancia_Menor = Distancia
            Ciudad_Mas_Cercana = Ciudad
   # print('remove '+str(Ciudad_Mas_Cercana))
    if Ciudad_Mas_Cercana != 0 :
        Ciudades_Disponibles.remove(Ciudad_Mas_Cercana)
    #print('Para '+str(Ciudad_Inicial)+' la ciudad más cercana es '+str(Ciudad_Mas_Cercana)+' con una distancia de '+str(Distancia_Menor)+' Kms')
    return Ciudad_Mas_Cercana

def Crear_Cromosomas_Iniciales(Cantidad_Cromosomas):
    Cromosomas = []
    Contador = 0
    while Contador < Cantidad_Cromosomas:
        Cromosoma = ''
        Ciudades_Disponibles = []
        for i in range(23):
            Ciudades_Disponibles.append(i+1)
        #print('Ciudades Disponibles ('+str(len(Ciudades_Disponibles))+'):')
        #print(Ciudades_Disponibles)
        while len(Ciudades_Disponibles) != 0:
            Posicion_Random = r.randint(0,len(Ciudades_Disponibles)-1)
            #print('Posicion_Random = '+str(Posicion_Random))
            Id_Ciudad_Random = Ciudades_Disponibles[Posicion_Random]
            #print('Id_Ciudad_Random = '+str(Id_Ciudad_Random))
            Ciudad_Random = Constant.CIUDADES[Id_Ciudad_Random-1]
            #print(Ciudad_Random)
            Gen = Ciudad_Random[0]
            #print('Gen = '+Gen)
            Cromosoma += Gen
            #print('cromosoma = '+cromosoma)
            Ciudades_Disponibles.pop(Posicion_Random)
            #print('')
            #print('Ciudades Disponibles ('+str(len(Ciudades_Disponibles))+'):')
            #print(Ciudades_Disponibles)
        if not(Cromosoma in Cromosomas):
            Cromosomas.append(Cromosoma)
            Contador+=1
    return Cromosomas
    #Contador = 0
    #for Cromosoma in Cromosomas:
        #Contador+=1
        #print(str(Contador)+' - '+cromosoma)

def Funcion_Objetivo(Cromosoma):
    distancia = 0
    ciudad_inicial = 0
    ciudad_final = 0
    for i in range(0,22):
        ciudad_actual = 0
        ciudad_proxima = 0
        for j in range(0, 23):
            if Cromosoma[i] == Constant.CIUDADES[j][0]:
                ciudad_actual = j+1
                if i == 0:
                    ciudad_inicial = j+1
                break
        for j in range(0, 23):
            if Cromosoma[i+1] == Constant.CIUDADES[j][0]:
                ciudad_proxima = j+1
                if i == 21:
                    ciudad_final = j+1
                break
        distancia += Sheet.cell(row=ciudad_actual,column=ciudad_proxima).value

    distancia += Sheet.cell(row=ciudad_inicial,column=ciudad_final).value
    return distancia

global ExcelDocument
global Sheet

ExcelDocument   = Excel.load_workbook(cwd+'\TP3\ '.strip() +FileName)  # Strip es para quitar los espacios en blanco, python no me deja poner la barra invertida al final de la cadena
Sheet           = ExcelDocument.get_sheet_by_name('PorRuta')

if __name__ == "__main__":
    main()
    